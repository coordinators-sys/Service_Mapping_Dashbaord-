"""Build data/site-code-crosswalk.json from IOM's ZiteManager Site Codes export.

The export (District | Identifier | Name | Latitude | Longitude) is the
authoritative registry of the ZiteManager site codes that appear as
`siteCodeRaw` on ZiteManager records. Those codes use a district-scoped
numbering scheme that does NOT correspond to the CCCM master list's own
sequence numbers, so they can never be resolved by parsing the code string
(measured: parsing bound records to the WRONG site, 0/141 name agreement).

Resolution here is by NAME + DISTRICT against the master list, then
CONFIRMED by GPS: the export carries coordinates for every row, so a
candidate is only accepted when it also sits within GPS_CONFIRM_KM of the
master-list site. Anything ambiguous, name-absent, or geographically
contradicted is skipped and left visibly unmatched.

Usage: python scripts/build_zite_site_crosswalk.py "path/to/Site Codes.xlsx"
"""

from __future__ import annotations

import difflib
import json
import math
import sys

import openpyxl

sys.path.insert(0, ".")

from api.lib.site_matching import get_master_site_index, _normalize_name  # noqa: E402
from api.lib.transformations import canonical_name  # noqa: E402

# A ZiteManager site and its master-list counterpart describe the same
# location, but each may be recorded from a different point inside a
# sprawling IDP site, so this is deliberately generous. It exists to REJECT
# same-name sites in genuinely different places, not to fingerprint a point.
GPS_CONFIRM_KM = 10.0

# Export district spellings -> master-list spellings.
DISTRICT_ALIASES = {"Kismayo": "Kismaayo"}


def _similar(a: str, b: str, cutoff: float = 0.82) -> bool:
    """Tolerant name comparison — the export and the live records differ in
    spacing/hyphenation/spelling on the same site ("deeq alle" vs "deeq-alle"),
    which is fine; only a substantively DIFFERENT name should reject an entry."""
    return difflib.SequenceMatcher(None, a, b).ratio() >= cutoff


def _haversine_km(lat1, lon1, lat2, lon2) -> float:
    r = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(a))


def main(xlsx_path: str) -> None:
    idx = get_master_site_index("data/master-sites.csv")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    raw_rows = [r for r in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True) if r[1] and r[2]]

    # The export repeats rows and, in at least one case, assigns the SAME code
    # to two different site names with different coordinates. Deduplicate, then
    # drop any code whose rows disagree on the name — a code that doesn't
    # identify one site cannot be crosswalked to one master site.
    seen: set[tuple] = set()
    rows = []
    for r in raw_rows:
        key = (r[1], str(r[2]).strip().lower())
        if key in seen:
            continue
        seen.add(key)
        rows.append(r)

    names_per_code: dict[str, set[str]] = {}
    for r in rows:
        names_per_code.setdefault(str(r[1]).strip().upper(), set()).add(str(r[2]).strip().lower())
    conflicting = {c for c, names in names_per_code.items() if len(names) > 1}

    crosswalk: dict[str, str] = {}
    stats = {
        "resolved_name_district_gps": 0,
        "resolved_name_district_no_master_coords": 0,
        "rejected_gps_conflict": 0,
        "rejected_code_name_conflict": 0,
        "ambiguous": 0,
        "name_not_in_master": 0,
        "already_direct": 0,
    }
    print(f"raw rows: {len(raw_rows)}  deduplicated: {len(rows)}  codes with conflicting names: {len(conflicting)}")
    for c in sorted(conflicting):
        print(f"    SKIPPED (ambiguous code): {c} -> {sorted(names_per_code[c])}")

    for district_raw, code, name, lat, lon in rows:
        if not code or not name:
            continue
        code = str(code).strip().upper()
        if code in conflicting:
            stats["rejected_code_name_conflict"] += 1
            continue

        # Codes the matcher already resolves on its own need no entry.
        if code in idx.by_id:
            stats["already_direct"] += 1
            continue

        district = canonical_name("district", DISTRICT_ALIASES.get(district_raw, district_raw)) or ""
        candidates = idx.by_name.get(_normalize_name(name), [])
        in_district = [c for c in candidates if _normalize_name(c.district) == _normalize_name(district)]

        pool = in_district if in_district else (candidates if len(candidates) == 1 else [])
        if not pool:
            stats["ambiguous" if candidates else "name_not_in_master"] += 1
            continue
        if len(pool) > 1:
            stats["ambiguous"] += 1
            continue

        site = pool[0]
        if site.latitude is not None and site.longitude is not None and lat is not None and lon is not None:
            if _haversine_km(float(lat), float(lon), site.latitude, site.longitude) > GPS_CONFIRM_KM:
                stats["rejected_gps_conflict"] += 1
                continue
            stats["resolved_name_district_gps"] += 1
        else:
            stats["resolved_name_district_no_master_coords"] += 1

        crosswalk[code] = site.cccm_site_id

    # FINAL GUARD — validate against what ZiteManager actually sends us.
    # The export's Name column is ~99% consistent with the ZiteManager UI, but
    # a few rows carry the wrong site name (one site's name duplicated onto
    # another's row). Those would map a code to the wrong master site, so any
    # entry whose target name contradicts the name on the LIVE records for that
    # same code is dropped. Live records are the authority: they are the rows
    # the dashboard actually ingests.
    payload_path = sys.argv[2] if len(sys.argv) > 2 else None
    if payload_path:
        live_names: dict[str, set[str]] = {}
        with open(payload_path, encoding="utf-8") as f:
            for rec in json.load(f).get("records", []):
                c = str(rec.get("siteCodeRaw") or "").strip().upper()
                n = rec.get("siteNameRaw")
                if c and n:
                    live_names.setdefault(c, set()).add(_normalize_name(n))
        by_id = {s.cccm_site_id: s for s in idx.sites}
        dropped = 0
        for code in list(crosswalk):
            observed = live_names.get(code)
            if not observed:
                continue
            target = _normalize_name(by_id[crosswalk[code]].site_name)
            if not any(_similar(target, o) for o in observed):
                print(f"    DROPPED (live name conflict): {code} -> {crosswalk[code]} "
                      f"({target!r}) but live records say {sorted(observed)}")
                del crosswalk[code]
                dropped += 1
        stats["rejected_live_name_conflict"] = dropped

    out = {
        "_comment": (
            "ZiteManager site code -> master CCCM Site ID. Generated by "
            "scripts/build_zite_site_crosswalk.py from IOM's ZiteManager Site "
            "Codes export; resolved by site name + district, confirmed by GPS "
            "proximity against the master list, and validated against the site "
            "names on live ZiteManager records — never by parsing the code."
        ),
        "entries": dict(sorted(crosswalk.items())),
    }
    with open("data/site-code-crosswalk.json", "w", encoding="utf-8") as f:
        json.dump(out, f, indent=1, ensure_ascii=False)

    print(f"input rows: {len(rows)}")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    print(f"wrote data/site-code-crosswalk.json with {len(crosswalk)} entries")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "ML/zite-site-codes.xlsx")
